import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

ABAS_TRIMESTRES = ["1 Trimestre", "2 Trimestre", "3 Trimestre"]

PENALIDADES_PADRAO = [
    ("P", -1),
    ("F", -2),
]

# Colunas fixas antes das aulas
COLUNAS_FIXAS = ["Aluno", "Média1", "Média2", "Média3", "Média", "Nº"]
COL_INICIO_AULAS = len(COLUNAS_FIXAS) + 1  # coluna G = 7


# ---------------------------------------------------------------------------
# Helpers de nome de arquivo
# ---------------------------------------------------------------------------

def _parsear_turma(nome_turma):
    """
    Extrai (serie, turno, letra) de strings como:
      "3ª Série - Noite - A"
      "ET GN - 2ª série - Tarde - K"
    Usa os três últimos segmentos separados por ' - '.
    """
    partes = [p.strip() for p in nome_turma.split(" - ")]
    if len(partes) >= 3:
        return partes[-3], partes[-2], partes[-1]
    if len(partes) == 2:
        return partes[0], "", partes[1]
    return nome_turma, "", ""


def _limpar(texto):
    """Remove caracteres que não podem ir no nome de arquivo."""
    return "".join(c for c in texto if c.isalnum() or c in ("_", "-")).strip()


def gerar_nome_arquivo(turma_data, pasta="planilhas"):
    escola_curta = _limpar(turma_data["escola"].split()[0])
    serie, turno, letra = _parsear_turma(turma_data["turma"])
    serie_limpa = _limpar(serie.replace("ª", "").replace(" ", ""))
    nome = f"{escola_curta}_{serie_limpa}_{turno}_{letra}.xlsx"
    return os.path.join(pasta, nome)


# ---------------------------------------------------------------------------
# Helpers de data
# ---------------------------------------------------------------------------

def _gerar_datas(data_inicio, num_aulas, frequencia_semanal):
    """
    Gera uma lista de datas de aula a partir de data_inicio.
    Distribui as aulas nos dias úteis da semana conforme a frequência.
    Dias de aula: segunda(0)..sexta(4), os primeiros `frequencia_semanal` dias.
    """
    if isinstance(data_inicio, str):
        data_inicio = date.fromisoformat(data_inicio)

    dias_de_aula = list(range(frequencia_semanal))  # 0=seg, 1=ter, ...
    datas = []
    atual = data_inicio

    while len(datas) < num_aulas:
        if atual.weekday() in dias_de_aula:
            datas.append(atual)
        atual += timedelta(days=1)

    return datas


# ---------------------------------------------------------------------------
# Aba Penalidades
# ---------------------------------------------------------------------------

def _criar_aba_penalidades(wb):
    ws = wb.create_sheet("Penalidades")

    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Ocorrencia").font = header_font
    ws.cell(row=1, column=2, value="Penalidade").font = header_font

    for i, (ocorrencia, penalidade) in enumerate(PENALIDADES_PADRAO, start=2):
        ws.cell(row=i, column=1, value=ocorrencia)
        ws.cell(row=i, column=2, value=penalidade)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12

    return ws


# ---------------------------------------------------------------------------
# Aba de trimestre
# ---------------------------------------------------------------------------

def _colunas_aulas(num_aulas, avaliacoes):
    """
    Retorna lista de dicts descrevendo cada coluna de aula:
      [{"tipo": "nota"|"ocorrencias"|"atividade", "aula": int, "avaliacao": int}, ...]
    Intercala uma coluna 'Atividade' após o último par de cada avaliação.
    """
    if not avaliacoes:
        avaliacoes = [{}]

    aulas_por_aval = num_aulas // len(avaliacoes)
    resto = num_aulas % len(avaliacoes)

    colunas = []
    aula_global = 1

    for idx_aval, _ in enumerate(avaliacoes):
        n = aulas_por_aval + (1 if idx_aval < resto else 0)
        for i in range(n):
            colunas.append({"tipo": "nota", "aula": aula_global, "avaliacao": idx_aval + 1})
            colunas.append({"tipo": "ocorrencias", "aula": aula_global, "avaliacao": idx_aval + 1})
            aula_global += 1
        colunas.append({"tipo": "atividade", "aula": None, "avaliacao": idx_aval + 1})

    return colunas


def _criar_aba_trimestre(wb, nome_aba, turma_data, config):
    ws = wb.create_sheet(nome_aba)

    num_aulas = config.get("num_aulas", 0)
    data_inicio = config.get("data_inicio")
    freq = config.get("frequencia_semanal", 1)
    avaliacoes = config.get("avaliacoes", [])
    alunos = sorted(turma_data.get("alunos", []), key=lambda a: a["numero"])

    titulo = (
        f"{turma_data['turma']}  |  "
        f"{turma_data['disciplina']}  |  "
        f"{turma_data['escola']}"
    )

    # Gera datas se possível
    datas = []
    if data_inicio and num_aulas > 0:
        datas = _gerar_datas(data_inicio, num_aulas, freq)

    colunas_aula = _colunas_aulas(num_aulas, avaliacoes)

    # ------------------------------------------------------------------
    # Linha 1 — título + datas
    # ------------------------------------------------------------------
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True)

    # Uma data por par (Nota, Ocorrencias); posiciona na coluna "Nota" de cada aula
    aula_col = {}  # aula_numero → coluna da célula "Nota"
    for i, col_info in enumerate(colunas_aula):
        col = COL_INICIO_AULAS + i
        if col_info["tipo"] == "nota":
            aula_num = col_info["aula"]
            aula_col[aula_num] = col
            if aula_num <= len(datas):
                cel = ws.cell(row=1, column=col, value=datas[aula_num - 1])
                cel.number_format = "DD/MM"
                cel.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------------------
    # Linha 2 — tema da aula (vazio, professor preenche)
    # ------------------------------------------------------------------
    ws.cell(row=2, column=1, value="Tema:")

    for i, col_info in enumerate(colunas_aula):
        col = COL_INICIO_AULAS + i
        if col_info["tipo"] == "nota":
            ws.cell(row=2, column=col, value="")  # espaço para o professor

    # ------------------------------------------------------------------
    # Linha 3 — cabeçalhos
    # ------------------------------------------------------------------
    for i, cabecalho in enumerate(COLUNAS_FIXAS, start=1):
        cel = ws.cell(row=3, column=i, value=cabecalho)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal="center")

    for i, col_info in enumerate(colunas_aula):
        col = COL_INICIO_AULAS + i
        tipo = col_info["tipo"]
        if tipo == "nota":
            label = "Nota"
        elif tipo == "ocorrencias":
            label = "Ocorrencias"
        else:
            label = f"Atividade {col_info['avaliacao']}"
        cel = ws.cell(row=3, column=col, value=label)
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------------------
    # Linhas 4+ — alunos
    # ------------------------------------------------------------------
    for idx, aluno in enumerate(alunos):
        row = 4 + idx
        ws.cell(row=row, column=1, value=aluno["nome"])
        ws.cell(row=row, column=6, value=aluno["numero"])

    # ------------------------------------------------------------------
    # Larguras de coluna
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = 35
    for col in range(2, COL_INICIO_AULAS):
        ws.column_dimensions[get_column_letter(col)].width = 10
    for i in range(len(colunas_aula)):
        col = COL_INICIO_AULAS + i
        col_info = colunas_aula[i]
        w = 12 if col_info["tipo"] == "atividade" else 8
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def gerar_planilha(turma_data, config, pasta="planilhas"):
    """
    Gera uma planilha Excel de diário de classe.

    Args:
        turma_data: dict com chaves: escola, turma, disciplina, alunos
        config: dict com: num_aulas, data_inicio (YYYY-MM-DD),
                frequencia_semanal, avaliacoes (lista de dicts com valor_maximo)
        pasta: diretório de saída

    Returns:
        Caminho do arquivo gerado.
    """
    os.makedirs(pasta, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba padrão vazia

    for nome_aba in ABAS_TRIMESTRES:
        _criar_aba_trimestre(wb, nome_aba, turma_data, config)

    _criar_aba_penalidades(wb)

    caminho = gerar_nome_arquivo(turma_data, pasta)
    wb.save(caminho)
    print(f"Planilha gerada: {caminho}")
    return caminho
